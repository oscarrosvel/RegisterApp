# app.py
from __future__ import annotations
from datetime import date, datetime, time
from decimal import Decimal
import csv, io, os, re, unicodedata, json

from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from flask import (
    Flask, render_template, request, jsonify, session, redirect, url_for, make_response, flash
)
from flask_sqlalchemy import SQLAlchemy
    # pip install psycopg2-binary si no tienes el driver
from sqlalchemy import func, and_
from werkzeug.security import generate_password_hash, check_password_hash


# ---------------------------------------------------------------------
# Configuración de la app y DB
# ---------------------------------------------------------------------
def make_app():
    # Habilitamos la carpeta "static" para poder usar url_for('static', ...)
    app = Flask(__name__, template_folder="templates", static_folder="static")
    app.secret_key = os.getenv("APP_SECRET", "change-me")

    # === PostgreSQL (recomendado) ===
    PG_USER = os.getenv("PGUSER", "register_user")
    PG_PASS = os.getenv("PGPASS", "register_pass")
    PG_HOST = os.getenv("PGHOST", "localhost")
    PG_PORT = os.getenv("PGPORT", "5432")
    PG_DB   = os.getenv("PGDATABASE", "registerapp")
    DB_URI  = f"postgresql+psycopg2://{PG_USER}:{PG_PASS}@{PG_HOST}:{PG_PORT}/{PG_DB}"

    app.config["SQLALCHEMY_DATABASE_URI"] = DB_URI
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    db = SQLAlchemy(app)

    # -----------------------------------------------------------------
    # Modelos
    # -----------------------------------------------------------------
    def today_default(): return date.today()

    class RazonSocial(db.Model):
        __tablename__ = "tbl_razon_social"
        id = db.Column(db.Integer, primary_key=True)
        nombre_razon_social = db.Column(db.String(255), nullable=False)

    class Restaurante(db.Model):
        __tablename__ = "tbl_restaurante"
        id = db.Column(db.Integer, primary_key=True)
        nom_restaurante = db.Column(db.String(255), nullable=False)
        id_razon_social = db.Column(db.Integer, db.ForeignKey("tbl_razon_social.id"), nullable=False)

    class Rol(db.Model):
        __tablename__ = "tbl_roles"
        id = db.Column(db.Integer, primary_key=True)
        nom_rol = db.Column(db.String(100), nullable=False)

    class Usuario(db.Model):
        __tablename__ = "tbl_usuario"
        id = db.Column(db.Integer, primary_key=True)
        nom_usuario = db.Column(db.String(120), nullable=False, unique=True)
        password_hash = db.Column(db.String(255), nullable=False)
        id_rol = db.Column(db.Integer, db.ForeignKey("tbl_roles.id"), nullable=False)
        id_razon_social = db.Column(db.Integer, db.ForeignKey("tbl_razon_social.id"), nullable=False)
        id_restaurante = db.Column(db.Integer, db.ForeignKey("tbl_restaurante.id"), nullable=True)
        activo = db.Column(db.Boolean, nullable=False, default=True)
        def set_password(self, raw): self.password_hash = generate_password_hash(raw)
        def check_password(self, raw): return check_password_hash(self.password_hash, raw)

    # -------- Operativas
    class TempEquipos(db.Model):
        __tablename__ = "tbl_temp_equipos"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        tipo_de_equipo = db.Column(db.String(50), nullable=False)
        num_equipo = db.Column(db.Integer, nullable=False)
        tipo_toma = db.Column(db.String(20), nullable=False)
        temperatura = db.Column(db.Numeric(6,2), nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    class TempAlimentos(db.Model):
        __tablename__ = "tbl_temp_alimentos"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        producto = db.Column(db.String(255), nullable=False)
        temperatura = db.Column(db.String(50), nullable=False)
        tiempo_preparacion = db.Column(db.Numeric(10,2), nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    class AceiteQuemado(db.Model):
        __tablename__ = "tbl_aceite_quemado"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        num_freidora = db.Column(db.Integer, nullable=False)
        filtracion = db.Column(db.Boolean, nullable=False)
        cambio_de_aceite = db.Column(db.Boolean, nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    class LimpiezaTrampasTanque(db.Model):
        __tablename__ = "tbl_limpieza_trampas_tanque"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        tipo_limpieza = db.Column(db.String(50), nullable=False)
        limpieza = db.Column(db.Boolean, nullable=False)
        desinfeccion = db.Column(db.Boolean, nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    class BPM(db.Model):
        __tablename__ = "tbl_bpm"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        nombre_auxiliar = db.Column(db.String(255), nullable=False)
        barba_maquillaje = db.Column(db.Boolean, nullable=False)
        cabello_gorro = db.Column(db.Boolean, nullable=False)
        ausencia_heridas = db.Column(db.Boolean, nullable=False)
        joyas_accesorios = db.Column(db.Boolean, nullable=False)
        perfumes = db.Column(db.Boolean, nullable=False)
        unas_manos = db.Column(db.Boolean, nullable=False)
        uniforme = db.Column(db.Boolean, nullable=False)
        zapatos = db.Column(db.Boolean, nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    class RecepcionMP(db.Model):
        __tablename__ = "tbl_recepcion_materias_primas"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        mp_insumo = db.Column(db.String(255), nullable=False)
        proveedor = db.Column(db.String(255), nullable=False)
        cantidad = db.Column(db.Numeric(12,2), nullable=False)
        temperatura = db.Column(db.String(50), nullable=False)
        lote = db.Column(db.String(100), nullable=False)
        fecha_vencimiento = db.Column(db.Date, nullable=False)
        n_factura = db.Column(db.String(100), nullable=False)
        requiere_cer_calidad = db.Column(db.Boolean, nullable=False)
        aceptado = db.Column(db.Boolean, nullable=False)
        transporte_limpio = db.Column(db.Boolean, nullable=False)
        termoking = db.Column(db.String(10), nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    class LimpiezaZonasCom(db.Model):
        __tablename__ = "tbl_limpieza_zonascom"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        establecimiento = db.Column(db.String(255), nullable=False)
        baño_hombre = db.Column(db.Boolean, nullable=False)
        baño_mujer = db.Column(db.Boolean, nullable=False)
        salon = db.Column(db.Boolean, nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    class LimpiezaGeneral(db.Model):
        __tablename__ = "tbl_limpieza_general"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        zona = db.Column(db.String(255), nullable=False)  # Zona/Equipo/Utensilio
        profunda = db.Column(db.Boolean, nullable=False)
        rutinaria = db.Column(db.Boolean, nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    class LimpiezaAlimentos(db.Model):
        __tablename__ = "tbl_limpieza_alimentos"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        alimento = db.Column(db.String(255), nullable=False)
        tiempo_exposicion = db.Column(db.String(30), nullable=False)
        tipo_desinfeccion = db.Column(db.String(30), nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    class AguaPotable(db.Model):
        __tablename__ = "tbl_agua_potable"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        olor = db.Column(db.Boolean, nullable=False)
        sabor = db.Column(db.Boolean, nullable=False)
        color = db.Column(db.Boolean, nullable=False)
        cloro = db.Column(db.Numeric(5,2), nullable=False)
        ph = db.Column(db.Numeric(5,2), nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    class ResiduosSolidos(db.Model):
        __tablename__ = "tbl_residuos_solidos"
        id = db.Column(db.Integer, primary_key=True)
        fecha = db.Column(db.Date, nullable=False, default=today_default)
        hora_disposicion_residuo = db.Column(db.Time, nullable=False)
        correcta_clasificacion = db.Column(db.Boolean, nullable=False)
        organico = db.Column(db.Boolean, nullable=False)
        reciclaje = db.Column(db.Boolean, nullable=False)
        ordinario = db.Column(db.Boolean, nullable=False)
        responsable = db.Column(db.String(255), nullable=False)
        observaciones = db.Column(db.Text)
        usuario = db.Column(db.String(50))

    # --- Conf Parametro Operativo
    class ConfParametroOperativo(db.Model):
        __tablename__ = "conf_parametro_operativo"
        id = db.Column(db.Integer, primary_key=True)
        tabla = db.Column(db.String(64), nullable=False, unique=True)  # nombre tabla operativa
        texto_html = db.Column(db.Text, nullable=False)                # HTML permitido (negrita, listas…)
        activo = db.Column(db.Boolean, nullable=False, default=True)
        actualizado = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow)
        
    class PermisosRol(db.Model):
        __tablename__ = "conf_permisos_rol"
        id = db.Column(db.Integer, primary_key=True)
        rol = db.Column(db.String(64), unique=True, nullable=False)
        tabs_json = db.Column(db.Text, nullable=False)  # lista JSON de keys de pestañas

    # Mapas
    MODEL_MAP = {
        "tbl_razon_social": RazonSocial,
        "tbl_restaurante": Restaurante,
        "tbl_roles": Rol,
        "tbl_usuario": Usuario,

        "tbl_temp_equipos": TempEquipos,
        "tbl_temp_alimentos": TempAlimentos,
        "tbl_aceite_quemado": AceiteQuemado,
        "tbl_limpieza_trampas_tanque": LimpiezaTrampasTanque,
        "tbl_bpm": BPM,
        "tbl_recepcion_materias_primas": RecepcionMP,
        "tbl_limpieza_zonascom": LimpiezaZonasCom,
        "tbl_limpieza_general": LimpiezaGeneral,
        "tbl_limpieza_alimentos": LimpiezaAlimentos,
        "tbl_agua_potable": AguaPotable,
        "tbl_residuos_solidos": ResiduosSolidos,
        "conf_parametro_operativo": ConfParametroOperativo,
    }
    
    # --- Orden y etiquetas bonitas ---
    ORDER_BY_TABLE = {
        "tbl_temp_equipos": ["fecha","tipo_de_equipo","num_equipo","tipo_toma","temperatura","responsable","observaciones"],
        "tbl_temp_alimentos": ["fecha","producto","temperatura","tiempo_preparacion","responsable","observaciones"],
        "tbl_aceite_quemado": ["fecha","num_freidora","filtracion","cambio_de_aceite","responsable","observaciones"],
        "tbl_limpieza_trampas_tanque": ["fecha","tipo_limpieza","limpieza","desinfeccion","responsable","observaciones"],
        "tbl_bpm": ["fecha","nombre_auxiliar","barba_maquillaje","cabello_gorro","ausencia_heridas","joyas_accesorios","perfumes","unas_manos","uniforme","zapatos","responsable","observaciones"],
        "tbl_recepcion_materias_primas": ["fecha","mp_insumo","proveedor","cantidad","temperatura","lote","fecha_vencimiento","n_factura","requiere_cer_calidad","aceptado","transporte_limpio","termoking","responsable","observaciones"],
        "tbl_limpieza_zonascom": ["fecha","establecimiento","baño_hombre","baño_mujer","salon","responsable","observaciones"],
        "tbl_limpieza_general": ["fecha","zona","profunda","rutinaria","responsable","observaciones"],
        "tbl_limpieza_alimentos": ["fecha","alimento","tiempo_exposicion","tipo_desinfeccion","responsable","observaciones"],
        "tbl_agua_potable": ["fecha","olor","sabor","color","cloro","ph","responsable","observaciones"],
        "tbl_residuos_solidos": ["fecha","hora_disposicion_residuo","correcta_clasificacion","organico","reciclaje","ordinario","responsable","observaciones"],
    }
    NICE_LABEL = {
        "fecha": "Fecha",
        "tipo_de_equipo": "Tipo de equipo",
        "num_equipo": "Nº equipo",
        "tipo_toma": "Tipo toma",
        "temperatura": "Temperatura (°C)",
        "responsable": "Responsable",
        "observaciones": "Observaciones",
        "producto": "Producto",
        "tiempo_preparacion": "Tiempo preparación (min)",
        "num_freidora": "Nº freidora",
        "filtracion": "Filtración",
        "cambio_de_aceite": "Cambio de aceite",
        "tipo_limpieza": "Tipo de trampa",
        "limpieza": "Limpieza",
        "desinfeccion": "Desinfección",
        "nombre_auxiliar": "Nombre auxiliar",
        "barba_maquillaje": "Barba / Maquillaje",
        "cabello_gorro": "Cabello / Gorro",
        "ausencia_heridas": "Ausencia de heridas",
        "joyas_accesorios": "Joyas / Accesorios",
        "perfumes": "Perfumes",
        "unas_manos": "Uñas / Manos",
        "uniforme": "Uniforme",
        "zapatos": "Zapatos",
        "mp_insumo": "MP / Insumo",
        "proveedor": "Proveedor",
        "cantidad": "Cantidad",
        "lote": "Lote",
        "fecha_vencimiento": "Fecha vencimiento",
        "n_factura": "Nº factura",
        "requiere_cer_calidad": "Req. cert. calidad",
        "aceptado": "Aceptado",
        "transporte_limpio": "Transporte limpio",
        "termoking": "Termoking",
        "establecimiento": "Establecimiento",
        "baño_hombre": "Baño hombre",
        "baño_mujer": "Baño mujer",
        "salon": "Salón",
        "zona": "Zona/Equipo/Utensilio",
        "alimento": "Alimento",
        "tiempo_exposicion": "Tiempo exposición",
        "tipo_desinfeccion": "Tipo desinfección",
        "olor": "Olor",
        "sabor": "Sabor",
        "color": "Color",
        "cloro": "Cloro (ppm)",
        "ph": "pH",
        "hora_disposicion_residuo": "Hora disposición",
        "correcta_clasificacion": "Correcta clasificación",
        "organico": "Orgánico (verde)",
        "reciclaje": "Reciclaje (blanca)",
        "ordinario": "Ordinario (negra)",
    }
    
    DEFAULT_ROLES = ["Admin","Supervisor","Operativo"]

    # --- Helpers dinámicos de permisos por rol
    def all_tab_keys():
        catalogs = ["tbl_razon_social","tbl_restaurante","tbl_roles","tbl_usuario","conf_parametro_operativo"]
        operatives = [
            "tbl_temp_equipos","tbl_temp_alimentos","tbl_aceite_quemado",
            "tbl_limpieza_trampas_tanque","tbl_bpm","tbl_recepcion_materias_primas",
            "tbl_limpieza_zonascom","tbl_limpieza_general","tbl_limpieza_alimentos",
            "tbl_agua_potable","tbl_residuos_solidos"
        ]
        others = ["tab_reportes","tab_permisos_roles"]
        return catalogs + operatives + others

    def default_tabs_for_role(role_name: str):
        """Defaults por rol si no hay config guardada en DB."""
        tabs_all = set(all_tab_keys())
        catalogs = {"tbl_razon_social","tbl_restaurante","tbl_roles","tbl_usuario","conf_parametro_operativo"}
        operatives = {
            "tbl_temp_equipos","tbl_temp_alimentos","tbl_aceite_quemado",
            "tbl_limpieza_trampas_tanque","tbl_bpm","tbl_recepcion_materias_primas",
            "tbl_limpieza_zonascom","tbl_limpieza_general","tbl_limpieza_alimentos",
            "tbl_agua_potable","tbl_residuos_solidos"
        }
        if role_name == "Admin":
            out = tabs_all.copy()
            out.add("tab_permisos_roles")  # Admin siempre puede ver la pestaña de permisos
            return sorted(out)
        else:
            out = set(operatives)
            out.add("tab_reportes")
            return sorted(out)

    def load_roles_tabs_from_db():
        """Devuelve un dict {rol: [tabs]} para TODOS los roles que existan en tbl_roles."""
        roles_in_db = [r.nom_rol for r in Rol.query.order_by(Rol.nom_rol).all()]
        rows = PermisosRol.query.all()
        saved = {r.rol: (json.loads(r.tabs_json) if r.tabs_json else []) for r in rows}

        matrix = {}
        for role in roles_in_db:
            if role in saved and saved[role]:
                tabs = set(saved[role])
            else:
                tabs = set(default_tabs_for_role(role))
            # Garantiza que Admin conserve la pestaña de permisos
            if role == "Admin":
                tabs.add("tab_permisos_roles")
            matrix[role] = sorted(tabs)
        return matrix

    FORMAL_NAMES = {
        "tbl_temp_equipos":"Temp. Equipos",
        "tbl_temp_alimentos":"Temp. Alimentos",
        "tbl_aceite_quemado":"Aceite Quemado",
        "tbl_limpieza_trampas_tanque":"Limpieza Trampas/Tanque",
        "tbl_bpm":"BPM",
        "tbl_recepcion_materias_primas":"Recepción M.P.",
        "tbl_limpieza_zonascom":"Limpieza Zonas Comunes",
        "tbl_limpieza_general":"Limpieza General",
        "tbl_limpieza_alimentos":"Limpieza Alimentos",
        "tbl_agua_potable":"Agua Potable",
        "tbl_residuos_solidos":"Residuos Sólidos",
    }

    def columns_of(model): return [c.name for c in model.__table__.columns]
    TABLES_CFG = {k: columns_of(v) for k, v in MODEL_MAP.items()}

    # -----------------------------------------------------------------
    # Bootstrap
    # -----------------------------------------------------------------
    with app.app_context():
        try:
            db.create_all()
        except Exception as e:
            print("\n[ERROR] No se pudo crear/esquema en PostgreSQL:", repr(e), "\n")
            raise

        # semillas mínimas
        if not Rol.query.first():
            db.session.add_all([Rol(nom_rol="Admin"), Rol(nom_rol="Supervisor"), Rol(nom_rol="Operativo")])
            db.session.commit()
        if not RazonSocial.query.first():
            rs = RazonSocial(nombre_razon_social="Inversiones alquimista")
            db.session.add(rs); db.session.commit()
        if not Restaurante.query.first():
            db.session.add(Restaurante(nom_restaurante="Alquimista", id_razon_social=RazonSocial.query.first().id))
            db.session.commit()
        if not Usuario.query.filter_by(nom_usuario="admin").first():
            u = Usuario(
                nom_usuario="admin",
                id_rol=Rol.query.filter_by(nom_rol="Admin").first().id,
                id_razon_social=RazonSocial.query.first().id,
                id_restaurante=Restaurante.query.first().id,
                activo=True
            )
            u.set_password("admin")
            db.session.add(u); db.session.commit()

    # -----------------------------------------------------------------
    # Helpers
    # -----------------------------------------------------------------
    def to_dict(obj):
        data = {}
        for c in obj.__table__.columns:
            v = getattr(obj, c.name)
            if isinstance(v, (date, datetime)):
                data[c.name] = v.isoformat()
            elif isinstance(v, time):
                data[c.name] = v.strftime("%H:%M:%S")
            elif isinstance(v, Decimal):
                data[c.name] = float(v)
            else:
                data[c.name] = v
        return data

    def parse_incoming(model, payload: dict):
        mapper = {c.name: c.type for c in model.__table__.columns}
        out = {}
        for k, v in (payload or {}).items():
            if k not in mapper: 
                continue
            if v in (None, ""):
                out[k] = None
                continue
            py = mapper[k].python_type
            if py is bool:
                if isinstance(v, bool): out[k] = v
                else: out[k] = str(v).lower() in ("true","1","si","sí","t","yes")
            elif py is int: out[k] = int(v)
            elif py is float: out[k] = float(v)
            elif py is Decimal: out[k] = Decimal(str(v))
            elif py is date: out[k] = datetime.strptime(v, "%Y-%m-%d").date()
            elif py is time:
                vv = v if len(v) > 5 else v + ":00"
                out[k] = datetime.strptime(vv, "%H:%M:%S").time()
            else: out[k] = v
        return out

    def slugify(value: str) -> str:
        """Convierte 'Inversiones Alquimista S.A.' -> 'inversiones-alquimista-s-a'"""
        value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
        value = re.sub(r'[^a-zA-Z0-9]+', '-', value).strip('-').lower()
        return value or 'logo'

    # -----------------------------------------------------------------
    # Auth mínima (ya tienes login.html propio)
    # -----------------------------------------------------------------
    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            ident = request.form.get("identifier")
            pwd = request.form.get("password")
            u = Usuario.query.filter_by(nom_usuario=ident).first()
            if u and u.check_password(pwd):
                session["usuario"] = u.nom_usuario
                session["rol"] = Rol.query.get(u.id_rol).nom_rol
                return redirect(url_for("register"))
            flash("Credenciales inválidas")
        return render_template("login.html")

    @app.route("/logout")
    def logout(): session.clear() or redirect(url_for("login")); return redirect(url_for("login"))

    # -----------------------------------------------------------------
    # UI principal
    # -----------------------------------------------------------------
    def allowed_tabs_for_role(rol: str):
        matrix = load_roles_tabs_from_db()
        return sorted(set(matrix.get(rol, default_tabs_for_role(rol))))

    @app.route("/register")
    def register():
        if "usuario" not in session:
            # Autologin de cortesía si entras directo
            user = Usuario.query.filter_by(nom_usuario="admin").first()
            session["usuario"] = user.nom_usuario
            session["rol"] = Rol.query.get(user.id_rol).nom_rol

        rol = session.get("rol", "Admin")

        # Catálogos
        razones = [ {"id":r.id,"nombre":r.nombre_razon_social} for r in RazonSocial.query.order_by(RazonSocial.nombre_razon_social).all() ]
        roles =   [ {"id":x.id,"nombre":x.nom_rol} for x in Rol.query.order_by(Rol.nom_rol).all() ]
        restaurantes = [{"id":r.id,"id_razon_social":r.id_razon_social,"nombre":r.nom_restaurante}
                        for r in Restaurante.query.order_by(Restaurante.nom_restaurante).all()]

        # Usuario actual y su razón social
        current_user = Usuario.query.filter_by(nom_usuario=session["usuario"]).first()
        rs_name = None
        if current_user and current_user.id_razon_social:
            rs = RazonSocial.query.get(current_user.id_razon_social)
            rs_name = rs.nombre_razon_social if rs else None

        # Construye logo dinámico (static/logos/<slug>.png) con fallback a default.png
        logo_url = None
        if rs_name:
            fname = f"logos/{slugify(rs_name)}.png"
            candidate_path = os.path.join(app.static_folder, fname)
            if os.path.exists(candidate_path):
                logo_url = url_for('static', filename=fname)
            else:
                # fallback opcional
                if os.path.exists(os.path.join(app.static_folder, "logos/default.png")):
                    logo_url = url_for('static', filename="logos/default.png")

        allowed_tabs = allowed_tabs_for_role(rol)
        
        all_roles = [r.nom_rol for r in Rol.query.order_by(Rol.nom_rol).all()]
        roles_tabs = load_roles_tabs_from_db()

        return render_template(
            "register.html",
            rol=rol,
            razones=razones,
            roles=roles,
            restaurantes=restaurantes,
            tables_cfg=TABLES_CFG,
            allowed_tabs=allowed_tabs,
            formal_names=FORMAL_NAMES,
            company_name=rs_name,
            logo_url=logo_url,
            all_roles=all_roles,
            roles_tabs=roles_tabs,
            nice_labels=NICE_LABEL,
            order_by_table=ORDER_BY_TABLE
        )

    # -----------------------------------------------------------------
    # API CRUD genérica
    # -----------------------------------------------------------------
    @app.route("/api/<table>", methods=["POST"])
    def api_create(table):
        M = MODEL_MAP.get(table)
        if not M:
            return "Tabla desconocida", 404

        payload = request.get_json(force=True, silent=True) or {}

        # ---- manejo especial para usuarios: contraseña en texto plano -> hash
        pwd = None
        if table == "tbl_usuario":
            # admite 'contraseña' (con tilde) o 'contrasena'
            pwd = payload.pop("contraseña", None) or payload.pop("contrasena", None)

        data = parse_incoming(M, payload)

        # si el modelo tiene columna 'usuario' y no viene, inyecta el de sesión
        if "usuario" in M.__table__.columns and not data.get("usuario"):
            data["usuario"] = session.get("usuario")

        obj = M(**data)

        if table == "tbl_usuario":
            if not pwd:
                return "La contraseña es requerida", 400
            obj.set_password(pwd)

        db.session.add(obj)
        db.session.commit()
        return jsonify(to_dict(obj)), 201

    @app.route("/api/<table>/<int:pk>", methods=["PUT"])
    def api_update(table, pk):
        M = MODEL_MAP.get(table)
        if not M:
            return "Tabla desconocida", 404

        obj = M.query.get_or_404(pk)
        payload = request.get_json(force=True, silent=True) or {}

        # ---- si viene nueva contraseña, hashearla
        new_pwd = None
        if table == "tbl_usuario":
            new_pwd = payload.pop("contraseña", None) or payload.pop("contrasena", None)

        data = parse_incoming(M, payload)
        for k, v in data.items():
            setattr(obj, k, v)

        if table == "tbl_usuario" and new_pwd:
            obj.set_password(new_pwd)

        db.session.commit()
        return jsonify(to_dict(obj))

    @app.route("/api/<table>/<int:pk>", methods=["DELETE"])
    def api_delete(table, pk):
        M = MODEL_MAP.get(table)
        if not M: return "Tabla desconocida", 404
        obj = M.query.get_or_404(pk)
        db.session.delete(obj); db.session.commit()
        return "", 204

    # -----------------------------------------------------------------
    # Reportes / Exportar
    # -----------------------------------------------------------------
    def build_filters(M, payload):
        f = []
        df, dt = payload.get("date_from"), payload.get("date_to")
        if hasattr(M, "fecha"):
            if df: f.append(M.fecha >= datetime.strptime(df,"%Y-%m-%d").date())
            if dt: f.append(M.fecha <= datetime.strptime(dt,"%Y-%m-%d").date())
        cf = payload.get("column_filters") or {}
        for k, v in cf.items():
            if not hasattr(M, k): continue
            col = getattr(M, k)
            txt = str(v).strip()
            if txt.lower() in ("true","false"):
                f.append(col == (txt.lower()=="true"))
            else:
                f.append(func.cast(col, db.String).ilike(f"%{txt}%"))
        return f

    @app.route("/api/query", methods=["POST"])
    def api_query():
        p = request.get_json(force=True) or {}
        table = p.get("table")
        M = MODEL_MAP.get(table)
        if not M: return "Tabla desconocida", 404
        q = M.query
        flt = build_filters(M, p)
        if flt: q = q.filter(and_(*flt))
        rows = q.order_by(M.id.desc()).limit(int(p.get("limit", 500))).all()
        return jsonify([to_dict(x) for x in rows])

    @app.route("/api/export", methods=["POST"])
    def api_export():
        p = request.get_json(force=True) or {}
        table = p.get("table")
        M = MODEL_MAP.get(table)
        if not M:
            return "Tabla desconocida", 404

        # ----- Consulta con filtros -----
        q = M.query
        flt = build_filters(M, p)
        if flt:
            q = q.filter(and_(*flt))
        rows = q.order_by(M.id.desc()).all()

        # ----- Columnas visibles + orden común (global) -----
        hide = {"id","password_hash","usuario","id_razon_social","id_rol","id_restaurante"}
        base_cols = [c.name for c in M.__table__.columns if c.name not in hide]
        ordered = ORDER_BY_TABLE.get(table, base_cols)
        cols = [c for c in ordered if c in base_cols] or base_cols

        # ----- Excel -----
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"

        # ===== Título con nombre formal de la tabla =====
        title_text = FORMAL_NAMES.get(table, table)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        tcell = ws.cell(row=1, column=1)
        tcell.value = f"Exportación: {title_text}"
        tcell.font = Font(size=14, bold=True)
        tcell.alignment = Alignment(horizontal="center")

        # ===== Resumen de filtros aplicados (si los hay) =====
        def _pretty_bool(v: str):
            s = str(v).strip().lower()
            if s in ("true","1","t","si","sí","yes"): return "Sí"
            if s in ("false","0","f","no"): return "No"
            return None

        filters_lines = []
        df = (p.get("date_from") or "").strip()
        dt = (p.get("date_to") or "").strip()
        if df and dt:   filters_lines.append(f"Rango de fechas: {df} — {dt}")
        elif df:        filters_lines.append(f"Fecha desde: {df}")
        elif dt:        filters_lines.append(f"Fecha hasta: {dt}")

        cf = p.get("column_filters") or {}
        ordered_cf_keys = [c for c in ORDER_BY_TABLE.get(table, cf.keys()) if c in cf] or list(cf.keys())
        for k in ordered_cf_keys:
            raw = cf[k]
            lbl = NICE_LABEL.get(k, k.replace('_',' ').title())
            pb = _pretty_bool(raw)
            if pb is None and isinstance(raw, str) and raw.strip():
                filters_lines.append(f'{lbl}: contiene "{raw.strip()}"')
            else:
                filters_lines.append(f"{lbl}: {pb if pb is not None else raw}")

        row_idx = 2
        for line in filters_lines:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(cols))
            c = ws.cell(row=row_idx, column=1)
            c.value = f"Filtro: {line}"
            c.font = Font(italic=True)
            row_idx += 1

        # ===== Encabezados bonitos =====
        def nice_label(col): return NICE_LABEL.get(col, col.replace('_',' ').title())

        header_row = row_idx
        ws.append([nice_label(c) for c in cols])

        # ===== Filas de datos (booleans a Sí/No) =====
        for r in rows:
            d = to_dict(r)
            ws.append([("Sí" if d.get(c) is True else "No" if d.get(c) is False else d.get(c, "")) for c in cols])

        # Estilo encabezado + autofiltro + panes congelados (título + filtros arriba)
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(cols))}{header_row + max(len(rows),1)}"
        ws.freeze_panes = f"A{header_row+1}"

        # Ajuste de ancho
        for idx in range(1, len(cols) + 1):
            max_len = 0
            for i in range(1, ws.max_row + 1):
                v = ws.cell(row=i, column=idx).value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(10, max_len + 2), 50)

        # Respuesta .xlsx
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = make_response(bio.read())
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["Content-Disposition"] = f'attachment; filename="export_{table}.xlsx"'
        return resp

    # -----------------------------------------------------------------
    # Conf. parámetro operativo: obtener mensaje activo por tabla
    # -----------------------------------------------------------------
    @app.route("/api/cpo/message/<table>", methods=["GET"])
    def api_cpo_message(table):
        row = ConfParametroOperativo.query.filter_by(tabla=table, activo=True).first()
        if not row: return jsonify({"active": False})
        return jsonify({"active": True, "html": row.texto_html})
        
    from flask import abort

    def require_admin():
        if session.get("rol") != "Admin":
            abort(403, description="Solo Admin")

    @app.route("/api/roles_tabs", methods=["GET"])
    def api_roles_tabs_get():
        require_admin()
        return jsonify(load_roles_tabs_from_db())
        
    @app.route("/api/<table>", methods=["GET"])
    def api_list(table):
        M = MODEL_MAP.get(table)
        if not M:
            return "Tabla desconocida", 404
        try:
            limit = int(request.args.get("limit", 100))
        except:
            limit = 100
        rows = M.query.order_by(M.id.desc()).limit(limit).all()
        return jsonify([to_dict(x) for x in rows])

    @app.route("/api/roles_tabs", methods=["POST"])
    def api_roles_tabs_set():
        require_admin()
        payload = request.get_json(force=True) or {}

        valid_tabs = set(all_tab_keys())
        valid_roles = {r.nom_rol for r in Rol.query.all()}

        for role, tabs in (payload or {}).items():
            if role not in valid_roles:
                continue
            clean = sorted(set([t for t in (tabs or []) if t in valid_tabs]))
            if role == "Admin" and "tab_permisos_roles" not in clean:
                clean.append("tab_permisos_roles")

            row = PermisosRol.query.filter_by(rol=role).first()
            if not row:
                row = PermisosRol(rol=role, tabs_json=json.dumps(clean))
                db.session.add(row)
            else:
                row.tabs_json = json.dumps(clean)

        db.session.commit()
        return jsonify({"ok": True})

    # -----------------------------------------------------------------
    # Raíz
    # -----------------------------------------------------------------
    @app.route("/")
    def index(): return redirect(url_for("register"))

    return app, db


if __name__ == "__main__":
    app, db = make_app()
    app.run(debug=True)
